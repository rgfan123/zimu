#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
01-parse-bundles.py
===================
新旧礼包源文件合并调研的解析脚本（可复现）。

输入（仓库根目录）：
  - 大者国风上架品（内容详情）-202605更新(1).xlsx  sheet「第2页」（新文件，36 个礼包，带商品条码/税率/成本）
  - 京东商品编号.xlsx                             sheet「Sheet2」（旧文件，23 个礼包，无条码，格式错乱）
  - 京东商品编号.xlsx                             sheet「Sheet1/Sheet3/Sheet4」（EMG→单品名称字典，仅用于把旧文件
    的 EMG 组件解码成名称，供与新版组件对比）

输出（本目录）：
  - 01-parsed.json                              解析后的新旧原始结构 + EMG 字典 + 重叠明细 + 合并清单
  - 01-merged-components.csv                    全组件明细（EMG 缺失标「缺」，含回填来源）
  - stdout 汇总统计

================================================================================
解析规则：旧 Sheet2 的「左右嵌」结构（务必按此理解，避免丢组件 / 重复计）
================================================================================
Sheet2 共 202 行，实际是 3 段内容并排在同一张表里：

1) 左块（A=礼包名 / B=EMG 组件编码 / C=数量）：完整的旧礼包 BOM，共 23 个礼包。
   每个礼包 = 连续若干行（同一礼包名重复出现），行间用空行分隔；组件数量以 C 列数字为准，
   与 B 列 EMG 一一对应。这是「23 个旧礼包」的权威来源。

2) 中缝（D 列）：恒为空，是左右两块的分隔列。

3) 右块（E/F/G 或 E/H/I）：另一批被并排贴进来的组件数据，非左块礼包的组成部分——
   误把它们并入左块礼包会造成「重复计」。右块自身的结构也分两种贴法：
     a) E=礼包名、F=EMG、G=数量（仅「新年礼包800」行1-7、「新年礼包1」行9-20 使用）；
     b) E=EMG、H='0'（占位符，即票面所说「0 1 占位」）、I=数量（两段无名数据，行22-31、行33-41）。
   右块的分段边界判定：出现新的 E 列礼包名、列位组合变化（F/G → E/H/I）、或出现空行中断，
   即视为下一段。E/H/I 两段没有礼包名，只能识别为「无名遗留块」。

「0 1 占位」：H 列固定为字符串 '0'，疑似从别处粘贴来的占位标记，不参与解析。

新文件「第2页」的组件解析规则：
  - A 列（商品条码）或 B 列（商品名称）出现 → 新礼包开始；其后 C 列有值、但 A/B 为空的行
    是该礼包的组件子行（C=组件名、D=数量、E=EMG）。
  - 父行 C 单元格可能内嵌全部组件，格式有三种：① 多行重复（每行一个组件，重复行数=数量，
    如行2）；② 单行粘连（组件名+重量+*数量 连写，如行3）；③ 多行带 *数量 或 /*N盒（如行4、15）。
  - 行 196-202：无条码的尾巴，内容与行 123-129 的「原切精品牛肉大礼包6140g」完全重复（无 EMG），
    判定为新文件内部的重复条目，不单独建礼包。
  - 组件级数量以「*N」「/*N盒」或子行 D 列为准；无数量标记时按 1 计；同名单行重复出现时数量累加。
"""
import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(BASE, "..", "..", ".."))
NEW_XLSX = os.path.join(REPO, "大者国风上架品（内容详情）-202605更新(1).xlsx")
OLD_XLSX = os.path.join(REPO, "京东商品编号.xlsx")

EMG_RE = re.compile(r"^EMG\d+$", re.I)

# ----------------------------------------------------------------------------
# 通用归一化
# ----------------------------------------------------------------------------

def norm_str(s):
    """全半角、空白、大小写归一（用于名称比较）。"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"\s+", "", s)
    return s.strip()


def norm_bundle_name(name):
    """礼包名归一：去空白/全半角、去（BJ）后缀、去【京东配送】前缀、去品牌前缀。

    品牌前缀剥离仅用于“是否同一礼包”的判定；合并清单里仍保留原始名称。
    """
    s = norm_str(name)
    s = s.replace("【京东配送】", "")
    s = re.sub(r"[（(]?\s*BJ\s*[）)]?$", "", s, flags=re.I)
    s = re.sub(r"^(子牧|子枚|卓宸|卓辰|北京)", "", s)
    return s


# 已知错别字/异体字别名（同一 EMG 或同一单品，仅名称写法不同）
# 注意：不要全局替换「羔羊」→「羊」，会破坏「羔羊羊颈排」；只替换确定的词组。
# 后腿块系列统一归一到「带骨羊后腿块」；已带「带骨」时用「带骨带骨→带骨」清理。
COMP_ALIAS = {
    "羔羊羊颈排": "羊颈排",
    "羔羊后羊腿块": "带骨羊后腿块",
    "羔羊后腿块": "带骨羊后腿块",
    "带骨带骨": "带骨",
    "羊排快": "羊排块",
    "牛助排": "牛肋排",
    "牛肋排块": "牛肋排",
    "后退块": "后腿块",
    "后退": "后腿",
    "帶骨": "带骨",
    "后羊腿块": "后腿块",
    "卓辰": "卓宸",
    "精选羔羊肉卷": "羔羊肉卷",
}

COMP_STRIP_PREFIX = re.compile(
    r"^(子牧|子枚|卓宸|卓辰|北京|原切|精品|精选|内蒙|新西兰|澳洲|谷饲|东乡贡羊|和牛M|M[0-9-]+)")


def comp_key(name):
    """组件名 → 比较键：去品牌/卖点前缀、去括号、去重量/单位/数量后缀。

    仅用于“旧文件 EMG 解码名”与“新文件无 EMG 组件名”之间的名称级近似匹配；
    EMG 级匹配优先于本函数。
    """
    s = norm_str(name)
    s = re.sub(r"[（(][^）)]*[）)]", "", s)          # 去掉括号及内容
    changed = True                                  # 别名循环到稳定
    while changed:
        changed = False
        for k, v in COMP_ALIAS.items():
            if k in s:
                s = s.replace(k, v)
                changed = True
    while True:                                     # 前缀可叠加（原切精选…），循环剥除
        s2 = COMP_STRIP_PREFIX.sub("", s)
        if s2 == s:
            break
        s = s2
    s = re.sub(r"\d+(\.\d+)?\s*(kg|千克|g|克|盒|袋|份|斤)", "", s, flags=re.I)
    s = re.sub(r"(\*|x|X|×)?\d+$", "", s)           # 尾部纯数字（无单位重量/数量）
    s = re.sub(r"[\s\*xX×]+$", "", s)
    return s.strip()


# ----------------------------------------------------------------------------
# 新文件解析
# ----------------------------------------------------------------------------

QTY_RE = re.compile(r"/\*\s*(\d+)\s*盒|\*\s*(\d+)|(\d+)\s*盒|(\d+)\s*袋|(\d+)\s*份")


def parse_single_comp_text(e):
    """把一段组件文本切成 (名称, 数量) 列表。"""
    e2 = re.sub(r"(\d)g(?=[0-9])", r"\1g*", e)      # '牛肉馅500g1' → '500g*1'
    parts = []
    last = 0
    for m in QTY_RE.finditer(e2):
        name = e2[last:m.start()].strip()
        qty = next(int(g) for g in m.groups() if g is not None)
        if name:
            parts.append((name, qty))
        last = m.end()
    tail = e2[last:].strip()
    if tail:
        parts.append((tail, 1))
    if not parts:
        parts.append((e.strip(), 1))
    return parts


def parse_comp_cell(text):
    """C 单元格（可能多行、可能内嵌整包组件）→ (名称, 数量) 列表。

    多行中行首为 ( （ ; ； , ， 、 的视为上一行的续行（组件名内嵌换行的括号说明），
    不另起组件。
    """
    if text is None:
        return []
    lines = [ln.strip() for ln in str(text).replace("\r", "").split("\n") if ln.strip()]
    entries = []
    for ln in lines:
        if entries and re.match(r"^[（(;；，,、]", ln):
            entries[-1] = entries[-1] + " " + ln
        else:
            entries.append(ln)
    out = []
    for en in entries:
        out.extend(parse_single_comp_text(en))
    return out


def parse_new_file():
    wb = openpyxl.load_workbook(NEW_XLSX, data_only=True)
    ws = wb["第2页"]
    bundles = []
    cur = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        a, b, c, d, e, f, g, h = row[:8]
        has_header = (a is not None and str(a).strip().startswith("925")) or (
            b is not None and str(b).strip() and not str(b).strip().startswith("商品名称"))
        if (a is not None and str(a).strip()) or (b is not None and str(b).strip()):
            if a is not None and str(a).strip():
                if not re.match(r"^925\d+$", str(a).strip()):
                    # 非条码的 A 值（如表头）跳过
                    if str(a).strip() == "商品条码":
                        continue
            # 新礼包开始
            cur = {
                "barcode": str(a).strip() if a is not None else None,
                "name": str(b).strip() if b is not None else None,
                "tax": f,
                "cost": g,
                "note": str(h).strip() if h is not None else None,
                "row": row[0].row if hasattr(row[0], "row") else None,
                "comps": [],
            }
            bundles.append(cur)
        if cur is None:
            continue
        # 组件：父行 C 内嵌文本 + 子行 C
        if c is not None:
            for nm, qty in parse_comp_cell(c):
                # 若单组件且文本未带显式数量，则用 D 列数量
                use_qty = qty
                if qty == 1 and d is not None and "*" not in str(c) and "/" not in str(c):
                    use_qty = int(d) if isinstance(d, (int, float)) else (int(str(d).strip()) if str(d).strip().isdigit() else 1)
                cur["comps"].append({
                    "name": nm,
                    "qty": use_qty,
                    "emg": str(e).strip() if e is not None else None,
                    "src": "parent" if b is not None else "child",
                })
    return bundles


# ----------------------------------------------------------------------------
# 旧文件 Sheet2 解析
# ----------------------------------------------------------------------------

def looks_emg(v):
    return v is not None and bool(EMG_RE.match(str(v).strip()))


def parse_old_sheet2():
    wb = openpyxl.load_workbook(OLD_XLSX, data_only=True)
    ws = wb["Sheet2"]
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append(list(r[:9]))

    # ---- 左块：23 个礼包（A=名称 B=EMG C=数量） ----
    left = []
    cur = None
    for i, r in enumerate(rows, start=1):
        a, b, c = r[0], r[1], r[2]
        an = str(a).strip() if a is not None else ""
        if an and (cur is None or cur["name"] != an):
            # 同一礼包名在连续多行重复出现；名字变化才算新礼包
            cur = {"name": an, "row_start": i, "comps": [], "block": "left"}
            left.append(cur)
        if b is not None and str(b).strip() and cur is not None:
            qty = int(c) if isinstance(c, (int, float)) else (
                int(str(c).strip()) if c is not None and str(c).strip().isdigit() else 1)
            cur["comps"].append({"emg": str(b).strip(), "qty": qty, "name": None})

    # ---- 右块：独立贴入的组件段（E/F/G 或 E/H/I） ----
    right = []          # 每段: {name, layout, rows:[...], comps:[...]}
    cur_r = None
    prev_layout = None
    for i, r in enumerate(rows, start=1):
        a, b, c, d, e, f, g, h, ii = r
        right_vals = [v for v in (e, f, g, h, ii) if v is not None and str(v).strip()]
        if not right_vals:
            cur_r = None
            prev_layout = None
            continue
        # 布局判断
        if looks_emg(e) and h is not None and str(h).strip() == "0" and ii is not None:
            layout = "E/H/I"
            nm, emg, qty = None, str(e).strip(), (int(ii) if isinstance(ii, (int, float)) else int(str(ii).strip()))
        elif f is not None and str(f).strip() and g is not None:
            layout = "E/F/G"
            nm_raw = str(e).strip() if e is not None else None
            nm = nm_raw if (nm_raw and not looks_emg(nm_raw)) else None
            emg = str(f).strip()
            qty = int(g) if isinstance(g, (int, float)) else (int(str(g).strip()) if str(g).strip().isdigit() else 1)
        else:
            layout = "unknown"
            nm, emg, qty = None, None, None

        start_new = False
        if cur_r is None:
            start_new = True
        elif nm is not None and (cur_r["name"] is None or nm != cur_r["name"]):
            start_new = True          # E 列出现新礼包名
        elif layout != prev_layout:
            start_new = True          # 贴法变化 → 新段
        if start_new:
            cur_r = {"name": nm, "layout": layout, "rows": [], "comps": [], "block": "right"}
            right.append(cur_r)
        if emg and qty:
            cur_r["comps"].append({"emg": emg, "qty": qty, "name": None})
        cur_r["rows"].append(i)
        prev_layout = layout

    return left, right


# ----------------------------------------------------------------------------
# EMG → 名称 字典（解码旧文件组件）
# ----------------------------------------------------------------------------

def build_emg_dict():
    wb = openpyxl.load_workbook(OLD_XLSX, data_only=True)
    d = {}          # emg -> [names]
    # Sheet1: F=EMG, E=JD 名称（短名）
    for r in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        e, f = r[4], r[5]
        if looks_emg(f):
            d.setdefault(str(f).strip(), []).append(str(e).strip())
    # Sheet3: A=EMG, B=名称（易和天下单品）
    for r in wb["Sheet3"].iter_rows(min_row=2, values_only=True):
        a, b = r[0], r[1]
        if looks_emg(a):
            d.setdefault(str(a).strip(), []).append(str(b).strip())
    # Sheet4: A=EMG, B=名称（聚福宝单品）
    for r in wb["Sheet4"].iter_rows(min_row=2, values_only=True):
        a, b = r[0], r[1]
        if looks_emg(a):
            d.setdefault(str(a).strip(), []).append(str(b).strip())
    # 新文件：E=EMG, C=组件名（BOM 内名，优先）
    wb2 = openpyxl.load_workbook(NEW_XLSX, data_only=True)
    for r in wb2["第2页"].iter_rows(min_row=2, values_only=True):
        c, e = r[2], r[4]
        if looks_emg(e) and c is not None:
            for nm, _q in parse_comp_cell(c):
                d.setdefault(str(e).strip(), []).append(nm)
    # 去重、去空
    out = {}
    for k, v in d.items():
        seen = []
        for n in v:
            n = n.strip()
            if n and n not in seen:
                seen.append(n)
        out[k] = seen
    return out


def decode_name(emg, emg_dict):
    names = emg_dict.get(emg)
    if names:
        # 优先挑最短名称作为“通用名”（品牌前缀最少）
        return min(names, key=len)
    return None


# ----------------------------------------------------------------------------
# 组件级匹配 & 礼包级相似度
# ----------------------------------------------------------------------------

def aggregate_comps(comps):
    """同一礼包内按组件名聚合数量（重复行/重复文本 = 数量累加），EMG 冲突时记录。"""
    agg = OrderedDict()
    for c in comps:
        key = norm_str(c["name"])
        if key not in agg:
            agg[key] = {"name": c["name"], "qty": 0, "emgs": []}
        agg[key]["qty"] += int(c["qty"] or 1)
        if c.get("emg") and c["emg"] not in agg[key]["emgs"]:
            agg[key]["emgs"].append(c["emg"])
    return list(agg.values())


def new_comp_by_key(bundle_comps):
    """新礼包：组件名 → comp（用于与旧组件名称级匹配）"""
    return {comp_key(c["name"]): c for c in bundle_comps}


def comp_emgs(c):
    """兼容 raw（'emg' 单值）与聚合（'emgs' 列表）两种组件结构。"""
    if c.get("emgs"):
        return c["emgs"]
    if c.get("emg"):
        return [c["emg"]]
    return []


def match_old_to_new(old_comp, new_comp_by_key):
    """旧组件（EMG+数量+解码名）在新礼包里找对应：先 EMG，后名称。"""
    all_new_emgs = {e for c in new_comp_by_key.values() for e in comp_emgs(c)}
    if old_comp["emg"] in all_new_emgs:
        return ("emg", old_comp["emg"])
    nm = decode_name(old_comp["emg"], old_comp["_emg_dict"])
    if nm:
        k = comp_key(nm)
        if k in new_comp_by_key:
            return ("name", new_comp_by_key[k]["name"])
    return None


def bundle_similarity(old_bundle, new_bundle, emg_dict):
    """旧礼包 vs 新礼包：组件匹配率。返回 (matched_old, total_old, detail)。"""
    nb_comps = new_bundle.get("comps_agg") or new_bundle["comps"]
    new_by_key = new_comp_by_key(nb_comps)
    matched = []
    for oc in old_bundle["comps"]:
        oc["_emg_dict"] = emg_dict
        m = match_old_to_new(oc, new_by_key)
        if m:
            matched.append((oc["emg"], m[1], oc["qty"]))
    return len(matched), len(old_bundle["comps"]), matched


# ----------------------------------------------------------------------------
# 主流程
# ----------------------------------------------------------------------------

def main():
    print("==" * 40)
    print("解析新文件：", NEW_XLSX)
    new_bundles = parse_new_file()
    # 标记无条码尾巴（重复条目）
    for b in new_bundles:
        b["comps_agg"] = aggregate_comps(b["comps"])
    print(f"  新文件礼包数: {len(new_bundles)}")
    for b in new_bundles:
        emg_n = sum(1 for c in b["comps_agg"] if c["emgs"])
        print(f"    [{b['barcode'] or '无条码'}] {b['name']}  组件{len(b['comps_agg'])} (带EMG {emg_n})")

    print()
    print("==" * 40)
    print("解析旧文件 Sheet2")
    left, right = parse_old_sheet2()
    for b in left:
        b["comps_agg"] = aggregate_comps([{"name": f"EMG:{c['emg']}", "qty": c["qty"], "emg": c["emg"]} for c in b["comps"]])
    print(f"  左块（正式旧礼包）: {len(left)} 个")
    for b in left:
        print(f"    {b['name']}  组件{len(b['comps_agg'])} (rows {b['row_start']}-)")
    print(f"  右块（遗留/并排数据）: {len(right)} 段")
    for b in right:
        print(f"    [{b['name'] or '无名'}] layout={b['layout']} rows={b['rows']}  组件{len(b['comps'])}")

    print()
    print("==" * 40)
    print("EMG → 名称 字典")
    emg_dict = build_emg_dict()
    print(f"  字典条目: {len(emg_dict)}")
    # 旧文件全部 EMG 的去码率
    all_old_emg = sorted({c["emg"] for b in left for c in b["comps"]} |
                         {c["emg"] for b in right for c in b["comps"]})
    undecodable = [e for e in all_old_emg if not decode_name(e, emg_dict)]
    print(f"  旧文件出现 EMG 总数: {len(all_old_emg)}，无法解码名称: {len(undecodable)} -> {undecodable}")

    # 解码旧组件名称
    for b in left + right:
        for c in b["comps"]:
            c["decoded"] = decode_name(c["emg"], emg_dict)

    # ------------------------------------------------------------------
    # 礼包级匹配（保守判定：名称归一一致 或 组件相似度>=0.8 才算重叠）
    # ------------------------------------------------------------------
    new_by_norm = {}
    for b in new_bundles:
        new_by_norm.setdefault(norm_bundle_name(b["name"]), []).append(b)

    def weight_proximity(a, b):
        """两个礼包名里标称重量的接近程度（克）；无重量信息返回一个大数。"""
        wa = [int(x) for x in re.findall(r"(\d{3,6})g", a)]
        wb = [int(x) for x in re.findall(r"(\d{3,6})g", b)]
        if not wa or not wb:
            return 10 ** 9
        return min(abs(x - y) for x in wa for y in wb)

    print()
    print("==" * 40)
    print("新旧礼包匹配（名称精确 + 组件相似度）")
    overlaps = []          # (old, new, mode, score)
    for ob in left:
        ob_norm = norm_bundle_name(ob["name"])
        candidates = []
        if ob_norm in new_by_norm:
            for nb in new_by_norm[ob_norm]:
                candidates.append((nb, "name-exact", 1.0))
        scored = []
        for nb in new_bundles:
            m, t, detail = bundle_similarity(ob, nb, emg_dict)
            if t == 0:
                continue
            score = m / t
            prox = weight_proximity(ob_norm, norm_bundle_name(nb["name"]))
            if prox <= 200:
                score = min(1.0, score + 0.05)
            scored.append((score, prox, m, t, nb, detail))
        scored.sort(key=lambda x: (-x[0], x[1]))
        # 仅取最优候选：名称归一一致（score 1.0）优先；否则 组件相似>=0.85 且 标称重量差<=200g
        if not candidates:
            for score, prox, m, t, nb, detail in scored:
                if score >= 0.85 and prox <= 200:
                    candidates.append((nb, f"compsim {m}/{t}", score))
                    break
        for nb, mode, score in candidates:
            overlaps.append((ob, nb, mode, score))
            print(f"  旧「{ob['name']}」 ↔ 新「{nb['name']}」 [{nb['barcode']}] 方式={mode} 相似={score:.2f}")

    # ------------------------------------------------------------------
    # 组件级合并（含旧文件 EMG 回填）
    # ------------------------------------------------------------------
    def old_by_key_map(ob, emg_dict):
        """旧礼包组件按名称键聚合（同名多行 = 数量累加；同名异 EMG 视为歧义）。"""
        m = {}
        for oc in ob["comps"]:
            nm = decode_name(oc["emg"], emg_dict)
            if not nm:
                continue
            k = comp_key(nm)
            ent = m.setdefault(k, {"emg": oc["emg"], "qty": 0, "name": nm, "ambiguous": False})
            ent["qty"] += int(oc["qty"] or 1)
            if ent["emg"] != oc["emg"]:
                ent["ambiguous"] = True
        return m

    def comp_level_merge(nb, ob, emg_dict):
        """重叠礼包组件级合并：
        - 新组件带 EMG → 保留（新文件优先）；
        - 新组件缺 EMG 但旧文件有同名组件（名称键匹配，且旧侧 EMG 无歧义）→ 回填旧 EMG；
        - 数量不一致时回填 EMG 但记 note。"""
        old_map = old_by_key_map(ob, emg_dict)
        out = []
        for nc in nb["comps_agg"]:
            rec = {"name": nc["name"], "qty": nc["qty"], "emg": None,
                   "emg_src": "缺", "note": ""}
            n_emgs = comp_emgs(nc)
            if n_emgs:
                rec["emg"] = n_emgs[0]
                rec["emg_src"] = "新文件"
            else:
                ent = old_map.get(comp_key(nc["name"]))
                if ent and not ent["ambiguous"]:
                    rec["emg"] = ent["emg"]
                    rec["emg_src"] = "旧文件回填"
                    if ent["qty"] != int(nc["qty"]):
                        rec["note"] = f"数量差异: 旧{ent['qty']} vs 新{nc['qty']}"
            out.append(rec)
        return out

    def old_comps_to_merged(ob, emg_dict):
        out = []
        for oc in ob["comps"]:
            nm = decode_name(oc["emg"], emg_dict)
            out.append({
                "name": nm or f"EMG:{oc['emg']}(名称待补)",
                "qty": oc["qty"],
                "emg": oc["emg"],
                "emg_src": "旧文件",
                "note": "" if nm else "旧文件EMG无法解码名称",
            })
        return out

    # ------------------------------------------------------------------
    # 合并清单
    # ------------------------------------------------------------------
    merged = []
    used_new = set()
    used_old = set()
    for ob, nb, mode, score in overlaps:
        if nb["barcode"] in used_new:
            # 多旧→一新（如旧 10000g/10100g → 新 9900g）：并入同一条记录
            for m in merged:
                if m["bundle_id"] == nb["barcode"] and ob["name"] not in m["old_names"]:
                    m["old_names"].append(ob["name"])
                    m["notes"].append(
                        f"旧文件另有同系礼包「{ob['name']}」并入本礼包（相似{score:.2f}，"
                        f"组件结构接近，疑为同系迭代/新旧版本）")
            used_old.add(id(ob))
            continue
        used_new.add(nb["barcode"])
        used_old.add(id(ob))
        merged.append({
            "bundle_id": nb["barcode"],
            "name": nb["name"],
            "source": "both",
            "barcode": nb["barcode"],
            "tax": nb["tax"],
            "cost": nb["cost"],
            "old_names": [ob["name"]],
            "match_mode": mode,
            "match_score": round(score, 3),
            "comps": comp_level_merge(nb, ob, emg_dict),
            "notes": [],
        })

    # 仅新（含无条码尾巴的处理）
    for nb in new_bundles:
        if nb["barcode"] is not None and nb["barcode"] not in used_new:
            merged.append({
                "bundle_id": nb["barcode"],
                "name": nb["name"],
                "source": "new",
                "barcode": nb["barcode"],
                "tax": nb["tax"],
                "cost": nb["cost"],
                "old_names": [],
                "match_mode": "new-only",
                "match_score": None,
                "comps": [{
                    "name": c["name"], "qty": c["qty"],
                    "emg": c["emgs"][0] if c["emgs"] else None,
                    "emg_src": "新文件" if c["emgs"] else "缺",
                    "note": "",
                } for c in nb["comps_agg"]],
                "notes": [],
            })
        elif nb["barcode"] is None:
            print(f"\n  [注意] 无条码尾巴条目: {nb['name']}（与有码条目原切精品牛肉大礼包6140g 组件完全重复，不建礼包）")

    # 仅旧（左块）
    old_only = [ob for ob in left if id(ob) not in used_old]
    for i, ob in enumerate(old_only, start=1):
        merged.append({
            "bundle_id": f"BUNDLE-{i:02d}",
            "name": ob["name"],
            "source": "old",
            "barcode": None,
            "tax": None,
            "cost": None,
            "old_names": [ob["name"]],
            "match_mode": "old-only",
            "match_score": None,
            "comps": old_comps_to_merged(ob, emg_dict),
            "notes": ["无条码，需生成内部标识 BUNDLE-xx"],
        })

    # 右块：命名段（新年礼包800/新年礼包1）纳入；无名段与疑似错位复制段单独列出
    right_import = []
    right_skip = []
    for rb in right:
        if rb["name"]:
            right_import.append(rb)
        else:
            # 疑似错位复制检测：与任一左块礼包按 EMG 多重集重叠>=80%
            dup_of = None
            best = 0.0
            rb_set = {}
            for c in rb["comps"]:
                rb_set[c["emg"]] = rb_set.get(c["emg"], 0) + 1
            for ob in left:
                ob_set = {}
                for c in ob["comps"]:
                    ob_set[c["emg"]] = ob_set.get(c["emg"], 0) + 1
                inter = sum(min(rb_set.get(k, 0), v) for k, v in ob_set.items())
                ov = inter / max(len(rb["comps"]), 1)
                if ov > best:
                    best, dup_of = ov, ob["name"]
            right_skip.append({"name": None, "layout": rb["layout"], "rows": rb["rows"],
                               "n_comps": len(rb["comps"]), "suspected_dup_of": dup_of,
                               "overlap": round(best, 2)})

    n_old_right = len(merged)
    for j, rb in enumerate(right_import, start=1):
        merged.append({
            "bundle_id": f"BUNDLE-R{j:02d}",
            "name": rb["name"],
            "source": "old-right",
            "barcode": None,
            "tax": None,
            "cost": None,
            "old_names": [rb["name"]],
            "match_mode": "old-right-only",
            "match_score": None,
            "comps": old_comps_to_merged(rb, emg_dict),
            "notes": ["旧 Sheet2 右嵌区块（E/F/G 布局）命名礼包；无条码，需生成内部标识"],
        })

    # ------------------------------------------------------------------
    # 字典级 EMG 解析（跨来源按名称唯一映射；歧义则保留「缺」）
    # ------------------------------------------------------------------
    name2emg = {}
    for emg, names in emg_dict.items():
        for n in names:
            name2emg.setdefault(comp_key(n), set()).add(emg)
    n_dict = 0
    for m in merged:
        for c in m["comps"]:
            if c["emg_src"] == "缺":
                cand = name2emg.get(comp_key(c["name"]))
                if cand and len(cand) == 1:
                    c["emg"] = next(iter(cand))
                    c["emg_src"] = "字典回填"
                    n_dict += 1

    # ------------------------------------------------------------------
    # 统计
    # ------------------------------------------------------------------
    print()
    print("==" * 40)
    print("合并统计")
    n_both = sum(1 for m in merged if m["source"] == "both")
    n_new = sum(1 for m in merged if m["source"] == "new")
    n_old = sum(1 for m in merged if m["source"] == "old")
    n_right = sum(1 for m in merged if m["source"] == "old-right")
    print(f"  新文件礼包: {len(new_bundles)}（含无条码重复尾巴 1 条，不计）")
    print(f"  旧文件左块礼包: {len(left)}")
    print(f"  重叠礼包(新+旧): {n_both}")
    print(f"  仅新: {n_new}")
    print(f"  仅旧(左块): {n_old}")
    print(f"  仅旧(右块命名礼包): {n_right}")
    print(f"  合并后礼包总数: {len(merged)}")
    for rs in right_skip:
        print(f"  [右块跳过] 无名段 rows={rs['rows']} 组件{rs['n_comps']} "
              f"疑似错位复制于「{rs['suspected_dup_of']}」(重叠{rs['overlap']})" if rs["suspected_dup_of"]
              else f"  [右块跳过] 无名段 rows={rs['rows']} 组件{rs['n_comps']}")

    # 缺 EMG 统计（全部回填/解析后）
    total_comp = 0
    missing_emg = 0
    backfilled_old = 0
    backfilled_dict = 0
    for m in merged:
        for c in m["comps"]:
            total_comp += 1
            if c["emg_src"] == "缺":
                missing_emg += 1
            elif c["emg_src"] == "旧文件回填":
                backfilled_old += 1
            elif c["emg_src"] == "字典回填":
                backfilled_dict += 1
    print(f"  合并后组件行总数: {total_comp}，缺 EMG: {missing_emg} "
          f"({missing_emg/total_comp*100:.1f}%)")
    print(f"  EMG 来源：新文件自带 {total_comp - missing_emg - backfilled_old - backfilled_dict}，"
          f"旧文件回填 {backfilled_old}，字典回填 {backfilled_dict}")

    # 重叠组件差异明细（供报告引用；旧侧按名称键聚合数量）
    overlap_detail = []
    for ob, nb, mode, score in overlaps:
        old_map = old_by_key_map(ob, emg_dict)
        rows = []
        for nc in nb["comps_agg"]:
            row = {"new_name": nc["name"], "new_qty": nc["qty"],
                   "new_emg": ",".join(comp_emgs(nc)) or None,
                   "old_emg": None, "old_name": None, "old_qty": None, "status": "仅新"}
            n_emgs = comp_emgs(nc)
            ent = old_map.get(comp_key(nc["name"]))
            if n_emgs:
                if ent:
                    row.update(old_emg=ent["emg"], old_name=ent["name"], old_qty=ent["qty"])
                    row["status"] = "EMG一致" if ent["emg"] in n_emgs else "EMG不同"
                else:
                    row["status"] = "仅新(EMG)"
            else:
                if ent:
                    row.update(old_emg=ent["emg"], old_name=ent["name"], old_qty=ent["qty"])
                    row["status"] = "名称匹配"
                else:
                    row["status"] = "仅新"
            rows.append(row)
        # 旧侧独有组件（新侧没有对应）
        new_keys = {comp_key(c["name"]) for c in nb["comps_agg"]}
        for k, ent in old_map.items():
            if k not in new_keys:
                rows.append({"new_name": None, "new_qty": None, "new_emg": None,
                             "old_emg": ent["emg"], "old_name": ent["name"],
                             "old_qty": ent["qty"], "status": "仅旧"})
        overlap_detail.append({
            "old": ob["name"], "new": nb["name"], "new_barcode": nb["barcode"],
            "mode": mode, "score": round(score, 3), "comps": rows,
        })

    # 输出 JSON
    out = {
        "new_bundles": new_bundles,
        "old_left": left,
        "old_right": right,
        "right_skipped": right_skip,
        "emg_dict": emg_dict,
        "overlap_detail": overlap_detail,
        "merged": merged,
    }
    with open(os.path.join(BASE, "01-parsed.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n  原始 JSON 已写入 01-parsed.json")

    # CSV：全组件明细
    import csv
    csv_path = os.path.join(BASE, "01-merged-components.csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["礼包标识", "礼包名称", "来源", "条码", "组件名称", "数量", "EMG编码", "EMG状态", "备注"])
        for m in merged:
            for c in m["comps"]:
                status = "缺" if not c.get("emg") else ("有" if c["emg_src"] != "旧文件回填" else "旧文件回填")
                w.writerow([m["bundle_id"], m["name"], m["source"],
                            m.get("barcode") or m["bundle_id"],
                            c["name"], c["qty"], c.get("emg") or "", status, c.get("note", "")])
    print(f"  CSV 已写入 01-merged-components.csv")

if __name__ == "__main__":
    main()
