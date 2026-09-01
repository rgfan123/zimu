#!/usr/bin/env python3
"""端到端：用真实结算账单复算运费。

需要账单 xlsx。查找顺序：环境变量 INVOICES > 相对本文件的 ../input/invoices/。
基线：735 命中 / 3 不符（均 >40kg 超重件，脚本自身会告警）/ 15 无档位。
"""
import os, re, sys, glob, importlib.util

HERE = os.path.dirname(os.path.abspath(__file__))

def locate_script():
    if os.environ.get("JDFEE"):
        return os.path.abspath(os.environ["JDFEE"])
    for c in (os.path.join(HERE, "scripts", "jdfee.py"),
              os.path.expanduser("~/.claude/skills/jd-fee-calc/scripts/jdfee.py")):
        if os.path.isfile(c):
            return c
    sys.exit("✗ 找不到 jdfee.py。用 JDFEE=<路径> 指定。")

def locate_invoices():
    if os.environ.get("INVOICES"):
        return os.environ["INVOICES"]
    c = os.path.normpath(os.path.join(HERE, "..", "input", "invoices"))
    if os.path.isdir(c):
        return c
    sys.exit("✗ 找不到账单目录。用 INVOICES=<目录> 指定。")

S, D = locate_script(), locate_invoices()
print(f"被测脚本：{S}\n账单目录：{D}\n")

spec = importlib.util.spec_from_file_location("jdfee", S)
jdfee = importlib.util.module_from_spec(spec); spec.loader.exec_module(jdfee)
import openpyxl

DISC = {"05": 0.30, "06": 0.40, "07": 0.20}   # 月度促销折扣
ok = bad = skip = 0; errs = []
files = sorted(glob.glob(os.path.join(D, "*收派*.xlsx")))
if not files:
    sys.exit(f"✗ {D} 下没有收派服务费账单")

for f in files:
    m = re.search(r"2026-(\d\d)-01", f)
    if not m: continue
    wb = openpyxl.load_workbook(f, data_only=True)
    if "明细_1" not in wb.sheetnames: continue
    ws = wb["明细_1"]; h = [c.value for c in ws[1]]
    if not h or h[0] is None: continue
    i = {x: j for j, x in enumerate(h)}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[i["费用类型"]] != "生鲜特惠运费": continue
        prov = r[i["目的省"]]; dest = (prov or "") + (r[i["目的市"]] or "")
        t = jdfee.DB["city"].get(dest)
        if not t: skip += 1; continue
        disc = 0.0 if prov == "新疆" else DISC[m.group(1)]
        res = jdfee.calc(t["b"], t["s"], float(r[i["实际重量(kg)"]]),
                         float(r[i["实际体积(cm3)"]]), freight_discount=disc)
        real = float(r[i["结算金额"]])
        if abs(res["freight"] - real) < 0.011: ok += 1
        else:
            bad += 1
            if len(errs) < 5:
                errs.append((dest, r[i["计费重量"]], res["freight"], real))

print(f"  精确命中 {ok}   不符 {bad}   无档位 {skip}")
for e in errs:
    print(f"    {e[0]}  计费{e[1]}kg  模型{e[2]}  实付{e[3]}")
BASE_OK, BASE_BAD = 735, 3
if ok >= BASE_OK and bad <= BASE_BAD:
    print(f"\n  ✓ 达到基线（≥{BASE_OK} 命中 / ≤{BASE_BAD} 不符）"); sys.exit(0)
print(f"\n  ✗ 低于基线（应 ≥{BASE_OK} 命中 / ≤{BASE_BAD} 不符）"); sys.exit(1)
