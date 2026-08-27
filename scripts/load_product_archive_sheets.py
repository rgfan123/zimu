#!/usr/bin/env python3
"""把成本表「成品」sheet 全列灌进 app.product_archive_sheets（V63），生成幂等 SQL。

    python3 scripts/load_product_archive_sheets.py <A产品成本核算26.3.29.xlsx> [> load.sql]

设计取舍（与 docs/research/jufubao-mapping-archive-2026-08-27.md 一致）：

* **列序是这张表存在的理由**：fields 存成 jsonb 数组而不是对象（PostgreSQL 的 jsonb 对象不
  保证键序），数组下标 == 原表列序 A..AU。空单元格保留元素、value 记 null 以保位。
* **指纹失败关闭**：源文件 SHA-256 与固定值不符即拒绝生成——沿用权威目录 manifest 的治理纪律
  （见 docs/authoritative-jd-catalog.md），不接受另存版本或手工转录。
* **幂等**：ON CONFLICT (source_file_sha256, row_no) DO NOTHING，重复灌不产生第二行。
* **挂接零猜测**：只有「品名逐字节相同 + 规格一致 + 没有第二行成本行争抢同一 SKU」的三行写
  matched_*，其余 107 行 matched_* 留空等人工挂接。按 sku_code 反查 id，不硬编码生产主键。
* 列语义（2026-08-27 用户拍板）：AI 线下供货成本/份 = 成本（按份 = 500g 单袋）；
  AJ 售价 = 不含运费售价。本脚本只留档，不回写 skus.purchase_price / retail_price。
"""

import hashlib
import json
import sys

import openpyxl
from openpyxl.utils import get_column_letter

EXPECTED_SHA256 = "e185b33fb5e856e9bdc324d6f4af8278ffb6937db3b09c4405f849208c2c86e4"
SOURCE_FILE_NAME = "A产品成本核算26.3.29.xlsx"
SHEET = "成品"

# 唯一达到「精确且无争抢」的挂接：品名逐字节相同、规格一致、无第二行争抢。
# row70/row72 同样精确同规格，但被上一位调研按「一 SKU 多行争抢」规则挂起，留待人工裁决。
CONFIDENT_MATCHES = {54: "SKU-JD-000011", 55: "SKU-JD-000012", 74: "SKU-JD-000033"}


def cell_text(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return str(int(value)) if value == int(value) and abs(value) < 1e15 else repr(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def quote(text):
    return "'" + text.replace("'", "''") + "'"


def main():
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    path = sys.argv[1]
    sha = hashlib.sha256(open(path, "rb").read()).hexdigest()
    if sha != EXPECTED_SHA256:
        raise SystemExit(
            f"source drift: expected {EXPECTED_SHA256}, got {sha}\n"
            "源文件已变更：停止生成，另起变更审查，不得直接更换脚本中的指纹。"
        )

    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    headers = [
        (col, get_column_letter(col), " ".join(str(ws.cell(1, col).value).split())
         if ws.cell(1, col).value is not None else None)
        for col in range(1, ws.max_column + 1)
    ]
    # 表格正身 = 有表头的列覆盖的连续区间 A..AU（AK 在区间内无表头但列列有数据）。
    # AU 之后是表格右侧零散手工草稿格，不是表的列，另存 extra_cells。
    last_headed = max(col for col, _, name in headers if name)
    table, outside = headers[:last_headed], headers[last_headed:]

    print("BEGIN;")
    rows = 0
    for r in range(2, ws.max_row + 1):
        product_name = cell_text(ws.cell(r, 1).value)
        if not product_name:
            continue
        rows += 1
        fields = [
            {"column": letter, "name": name or f"（{letter} 列无表头）",
             "value": cell_text(ws.cell(r, col).value)}
            for col, letter, name in table
        ]
        extra = [
            {"column": letter, "value": cell_text(ws.cell(r, col).value)}
            for col, letter, _ in outside
            if cell_text(ws.cell(r, col).value) is not None
        ]
        sku_code = CONFIDENT_MATCHES.get(r)
        matched_sku = f"(SELECT id FROM app.skus WHERE sku_code = {quote(sku_code)})" if sku_code else "NULL"
        matched_product = (
            f"(SELECT product_id FROM app.skus WHERE sku_code = {quote(sku_code)})" if sku_code else "NULL"
        )
        print(
            "INSERT INTO app.product_archive_sheets ("
            "source_file_name, source_file_sha256, sheet_name, row_no, product_name, "
            "fields, extra_cells, matched_sku_id, matched_product_id) VALUES ("
            f"{quote(SOURCE_FILE_NAME)}, {quote(sha)}, {quote(SHEET)}, {r}, {quote(product_name)}, "
            f"{quote(json.dumps(fields, ensure_ascii=False))}::jsonb, "
            f"{quote(json.dumps(extra, ensure_ascii=False))}::jsonb, "
            f"{matched_sku}, {matched_product}) "
            "ON CONFLICT (source_file_sha256, row_no) DO NOTHING;"
        )

    print("COMMIT;")
    print(
        f"-- 期望：{rows} 行 × {len(table)} 列（A..{table[-1][1]}），"
        f"其中 {len(CONFIDENT_MATCHES)} 行带挂接。",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
